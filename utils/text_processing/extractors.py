"""
Text extraction utilities for the Knowledge Base upload pipeline.
Supports PDF, DOCX, TXT, CSV and Excel files.
"""
import logging
import os

from pypdf import PdfReader
from docx import Document

logger = logging.getLogger("brightside")


class UnsupportedFileTypeError(Exception):
    pass


def extract_text(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return _extract_pdf(file_path)
    if ext == ".docx":
        return _extract_docx(file_path)
    if ext == ".txt":
        return _extract_txt(file_path)
    if ext == ".csv":
        return _extract_csv(file_path)
    if ext == ".xlsx":
        return _extract_xlsx(file_path)
    if ext == ".xls":
        return _extract_xls(file_path)

    raise UnsupportedFileTypeError(f"Unsupported file type: {ext}")


def _extract_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx(file_path: str) -> str:
    document = Document(file_path)
    return "\n".join(p.text for p in document.paragraphs)


def _extract_txt(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _extract_csv(file_path: str) -> str:
    import csv
    lines = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            row_text = " | ".join(cell.strip() for cell in row if cell.strip())
            if row_text:
                lines.append(row_text)
    return "\n".join(lines)


def _extract_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True, read_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())
            if row_text:
                lines.append(row_text)
    return "\n".join(lines)


def _extract_xls(file_path: str) -> str:
    import xlrd
    wb = xlrd.open_workbook(file_path)
    lines = []
    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        lines.append(f"--- Sheet: {sheet.name} ---")
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            row_text = " | ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())
            if row_text:
                lines.append(row_text)
    return "\n".join(lines)
